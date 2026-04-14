"""Endpoints para la capa IP/ISP: ruteadores, proveedores, flujos de tráfico y simulación ISP."""
from collections import Counter
from flask import Blueprint, request, jsonify
from ..extensions import db
from ..models import Router, ISPProvider, RouterInterface, TrafficFlow, LambdaUtilization, Lambda

bp = Blueprint("isp", __name__)


# ── Helper: endpoint remoto de una lambda desde un sitio dado ─────────────────

def _lambda_remote_site(lambda_, local_site_id):
    """Devuelve el sitio del otro extremo de la lambda, o None si no se puede determinar."""
    counter = Counter()
    for ls in lambda_.lambda_segments:
        seg = ls.segment
        counter[seg.site_a_id] += 1
        counter[seg.site_b_id] += 1
    # Los sitios terminales aparecen una sola vez en la cadena de segmentos
    endpoints = [s for s, cnt in counter.items() if cnt == 1]
    if not endpoints and lambda_.lambda_segments:
        # Lambda de un solo segmento: ambos sitios son terminales
        seg = lambda_.lambda_segments[0].segment
        endpoints = [seg.site_a_id, seg.site_b_id]
    return next((s for s in endpoints if s != local_site_id), None)


# ── Ruteadores ────────────────────────────────────────────────────────────────

@bp.get("/routers")
def get_routers():
    routers = Router.query.all()
    return jsonify([r.to_dict(include_interfaces=True) for r in routers])


@bp.put("/router-interfaces/<int:iface_id>")
def update_router_interface(iface_id):
    """Actualiza la métrica ISIS de una interfaz lambda (Cisco/Juniper)."""
    iface = RouterInterface.query.get_or_404(iface_id)
    data  = request.get_json(silent=True) or {}

    if "isis_metric" in data:
        val = data["isis_metric"]
        if not isinstance(val, int) or not (1 <= val <= 16777214):
            return jsonify({"error": "isis_metric debe ser un entero entre 1 y 16 777 214"}), 422
        if iface.iface_type != "lambda":
            return jsonify({"error": "isis_metric solo aplica a interfaces tipo 'lambda'"}), 422
        iface.isis_metric = val

    db.session.commit()
    return jsonify(iface.to_dict())


# ── Proveedores ISP ───────────────────────────────────────────────────────────

@bp.get("/isp-providers")
def get_isp_providers():
    providers = ISPProvider.query.order_by(ISPProvider.name).all()
    return jsonify([p.to_dict() for p in providers])


# ── Flujos de tráfico ─────────────────────────────────────────────────────────

@bp.get("/traffic-flows")
def get_traffic_flows():
    flows = TrafficFlow.query.all()
    return jsonify([f.to_dict() for f in flows])


@bp.put("/traffic-flows/<int:flow_id>")
def update_traffic_flow(flow_id):
    flow = TrafficFlow.query.get_or_404(flow_id)
    data = request.get_json(silent=True) or {}

    if "interfaces_count" in data:
        count = int(data["interfaces_count"])
        if count < 0:
            return jsonify({"error": "interfaces_count no puede ser negativo"}), 400
        # Validar que no supere la capacidad ISP disponible en el sitio ingress
        isp_capacity = RouterInterface.query.filter_by(
            isp_provider_id=flow.isp_provider_id,
        ).join(Router).filter(Router.site_id == flow.ingress_site_id).count()
        if count > isp_capacity:
            return jsonify({
                "error": f"interfaces_count ({count}) supera la capacidad ISP "
                         f"disponible en el sitio ({isp_capacity} interfaces)"
            }), 400
        flow.interfaces_count = count

    db.session.commit()
    return jsonify(flow.to_dict())


# ── Simulación ISP ────────────────────────────────────────────────────────────

@bp.post("/simulation/isp-provider")
def simulate_isp_provider():
    """
    Simula la caída de un proveedor ISP en un sitio.
    Calcula redistribución proporcional local y, si queda déficit,
    sugiere rerouteo ISIS por interfaces lambda (métrica ascendente).
    """
    data          = request.get_json(silent=True) or {}
    provider_name = data.get("provider")
    site_id       = data.get("site_id")

    if not provider_name or not site_id:
        return jsonify({"error": "Se requieren 'provider' y 'site_id'"}), 400

    provider = ISPProvider.query.filter_by(name=provider_name).first()
    if not provider:
        return jsonify({"error": f"Proveedor '{provider_name}' no encontrado"}), 404

    router = Router.query.filter_by(site_id=site_id).first()
    if not router:
        return jsonify({"error": f"No hay ruteador en el sitio '{site_id}'"}), 404

    # ── Flujos e interfaces afectados ─────────────────────────────────────────

    affected_flows = TrafficFlow.query.filter_by(
        isp_provider_id=provider.id,
        ingress_site_id=site_id,
    ).all()

    affected_gbps  = sum(f.interfaces_count * 100 for f in affected_flows)
    affected_ifaces = sum(f.interfaces_count for f in affected_flows)

    provider_ifaces_total = RouterInterface.query.filter_by(
        router_id=router.id,
        iface_type="isp",
        isp_provider_id=provider.id,
    ).count()

    # ── Capacidad disponible de OTROS proveedores en el mismo sitio ───────────

    from ..models import ISPProvider as ISP
    other_providers = db.session.query(ISP).join(RouterInterface).filter(
        RouterInterface.router_id == router.id,
        RouterInterface.iface_type == "isp",
        RouterInterface.isp_provider_id != provider.id,
    ).distinct().all()

    other_summary = []
    for op in other_providers:
        op_ifaces_total = RouterInterface.query.filter_by(
            router_id=router.id, iface_type="isp", isp_provider_id=op.id
        ).count()
        op_used = sum(
            f.interfaces_count
            for f in TrafficFlow.query.filter_by(isp_provider_id=op.id, ingress_site_id=site_id).all()
        )
        op_avail = max(0, op_ifaces_total - op_used)
        other_summary.append({
            "provider": op.name,
            "color": op.color,
            "total_ifaces": op_ifaces_total,
            "used_ifaces": op_used,
            "available_ifaces": op_avail,
        })

    available_ifaces = sum(s["available_ifaces"] for s in other_summary)
    available_gbps   = available_ifaces * 100
    deficit_gbps     = max(0, affected_gbps - available_gbps)

    # ── Redistribución proporcional (local) ────────────────────────────────────
    # Protocolo IGP ISIS asumido; todas las interfaces ISP con la misma métrica
    # (sin peso diferencial entre proveedores ISP externos).

    redistribution_detail = []
    if available_ifaces > 0 and affected_ifaces > 0:
        for s in other_summary:
            if s["available_ifaces"] > 0:
                prop          = s["available_ifaces"] / available_ifaces
                absorbed_gbps = round(affected_gbps * prop, 1)
                redistribution_detail.append({
                    "provider":       s["provider"],
                    "color":          s["color"],
                    "interfaces":     s["available_ifaces"],
                    "capacity_gbps":  s["total_ifaces"] * 100,
                    "absorbed_gbps":  absorbed_gbps,
                    "overloaded":     absorbed_gbps > s["total_ifaces"] * 100,
                })

    # ── Rerouteo ISIS por interfaces lambda (solo si hay déficit) ────────────
    # Cuando la redistribución local no cubre el tráfico afectado, ISIS puede
    # reconverger y enrutar hacia sitios vecinos con conectividad ISP disponible.
    # Se ordenan por métrica ISIS ascendente (menor métrica = ruta preferida).

    isis_rerouting = []
    if deficit_gbps > 0 and router.brand in ("cisco", "juniper"):
        lambda_ifaces = [
            i for i in router.interfaces
            if i.iface_type == "lambda" and i.lambda_ and i.isis_metric is not None
        ]
        seen_remote = set()
        for iface in sorted(lambda_ifaces, key=lambda x: x.isis_metric):
            remote_site_id = _lambda_remote_site(iface.lambda_, site_id)
            if not remote_site_id or remote_site_id in seen_remote:
                continue
            seen_remote.add(remote_site_id)

            remote_router = Router.query.filter_by(site_id=remote_site_id).first()
            if not remote_router:
                continue

            # Capacidad ISP total en el sitio remoto
            remote_isp_total = RouterInterface.query.filter_by(
                router_id=remote_router.id, iface_type="isp"
            ).count() * 100  # Gbps

            # Tráfico ya configurado en el sitio remoto
            remote_used = sum(
                f.interfaces_count * 100
                for f in TrafficFlow.query.filter_by(ingress_site_id=remote_site_id).all()
            )
            remote_available = max(0, remote_isp_total - remote_used)

            isis_rerouting.append({
                "interface_name":          iface.name,
                "lambda_name":             iface.lambda_.name,
                "isis_metric":             iface.isis_metric,
                "remote_site_id":          remote_site_id,
                "remote_site_name":        remote_router.site.name if remote_router.site else remote_site_id,
                "remote_isp_capacity_gbps": remote_isp_total,
                "remote_isp_available_gbps": remote_available,
            })

    return jsonify({
        "provider":              provider_name,
        "provider_color":        provider.color,
        "site_id":               site_id,
        "site_name":             router.site.name if router.site else site_id,
        "provider_ifaces_total": provider_ifaces_total,
        "affected_flows_count":  len(affected_flows),
        "affected_gbps":         affected_gbps,
        "available_gbps":        available_gbps,
        "deficit_gbps":          deficit_gbps,
        "status":                "deficit" if deficit_gbps > 0 else ("redistributed" if affected_gbps > 0 else "no_traffic"),
        "redistribution_detail": redistribution_detail,
        "isis_rerouting":        isis_rerouting,
        "affected_flows":        [f.to_dict() for f in affected_flows],
    })


@bp.post("/simulation/lambda-traffic")
def simulate_lambda_traffic():
    """Calcula el impacto de la caída de una o varias lambdas sobre los flujos de tráfico."""
    data = request.get_json(silent=True) or {}
    lambda_names = data.get("lambda_names", [])  # lista de nombres de lambda

    if not lambda_names:
        return jsonify({"error": "Se requiere 'lambda_names' (lista de nombres)"}), 400

    failed_set = set(lambda_names)
    all_flows  = TrafficFlow.query.all()

    affected = []
    for flow in all_flows:
        if not flow.lambda_names or flow.interfaces_count == 0:
            continue
        flow_lambdas = {n.strip() for n in flow.lambda_names.split(",")}
        impacted = flow_lambdas & failed_set
        if impacted:
            affected.append({
                **flow.to_dict(),
                "failed_lambdas": list(impacted),
                "traffic_at_risk_gbps": flow.interfaces_count * 100,
            })

    total_at_risk = sum(a["traffic_at_risk_gbps"] for a in affected)

    return jsonify({
        "failed_lambdas": list(failed_set),
        "affected_flows_count": len(affected),
        "total_traffic_at_risk_gbps": total_at_risk,
        "affected_flows": affected,
    })


# ── Utilización mensual de lambdas (importación Excel) ───────────────────────

# Mapeo de nombres de sitio del Excel a Site IDs en la BD
_EXCEL_SITE_MAP = {
    "mso tlalnepantla": "MSOMEX01",
    "tlalnepantla":     "MSOMEX01",
    "mso megacentro":   "MSOMEX01",
    "megacentro":       "MSOMEX01",
    "mso apodaca":      "MSOMTY01",
    "apodaca":          "MSOMTY01",
    "mso monterrey":    "MSOMTY01",    # Apodaca es en Monterrey
    "mso tlaquepaque":  "MSOGDL01",
    "tlaquepaque":      "MSOGDL01",
    "mso guadalajara":  "MSOGDL01",
    "guadalajara":      "MSOGDL01",
    "mtx toluca":       "MSOTOL01",
    "toluca":           "MSOTOL01",
    "nuevo laredo":     "NFO-038",
    "laredo":           "NFO-038",
    "reynosa":          "TAMREY1273",
    "cd. juarez":       "MSOJRZ01",
    "cd juarez":        "MSOJRZ01",
    "juarez":           "MSOJRZ01",
    "kio queretaro":    "KIO-QRO",
    "kio qro":          "KIO-QRO",
    "kio tultitlan":    None,   # sitio desconocido
}


def _resolve_site(name_fragment: str):
    """Mapea un fragmento de nombre de sitio del Excel al Site ID correspondiente."""
    key = name_fragment.strip().lower()
    for k, v in _EXCEL_SITE_MAP.items():
        if k in key:
            return v
    return None  # desconocido


def _find_lambda_for_link(site_a_id, site_b_id):
    """Intenta encontrar una lambda cuyos endpoints sean site_a y site_b."""
    if not site_a_id or not site_b_id:
        return None
    all_lambdas = Lambda.query.all()
    for lm in all_lambdas:
        freq = Counter()
        for ls in lm.lambda_segments:
            seg = ls.segment
            freq[seg.site_a_id] += 1
            freq[seg.site_b_id] += 1
        endpoints = [s for s, c in freq.items() if c == 1]
        if not endpoints and lm.lambda_segments:
            seg = lm.lambda_segments[0].segment
            endpoints = [seg.site_a_id, seg.site_b_id]
        ep_set = set(endpoints)
        if site_a_id in ep_set and site_b_id in ep_set:
            return lm
    return None


@bp.post("/upload/lambda-utilization")
def upload_lambda_utilization():
    """
    Importa datos de utilización mensual desde un archivo Excel.
    Multipart: campo 'file' = archivo .xlsx
    Formato esperado: hoja THP_Marzo26
      Fila 0: fechas de mes (col 4, 8, 12, 16, 20, 24)
      Fila 1: cabeceras (Enlace, Equipo A, Equipo B, BW Gb, Max Gbps, % Util Max, AVG Gbps, % Util AVG, ...)
      Filas 2+: datos
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl no está instalado"}), 500

    if "file" not in request.files:
        return jsonify({"error": "Se requiere el campo 'file'"}), 400

    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "El archivo debe ser .xlsx o .xls"}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({"error": f"No se pudo abrir el archivo: {e}"}), 400

    # Buscar la hoja correcta (THP_Marzo26 o la activa)
    ws = wb["THP_Marzo26"] if "THP_Marzo26" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 3:
        return jsonify({"error": "El archivo no tiene el formato esperado"}), 400

    # Extraer fechas de mes (fila 0, columnas 4,8,12,16,20,24)
    month_row = rows[0]
    month_cols = [4, 8, 12, 16, 20, 24]
    months = []
    for col_idx in month_cols:
        val = month_row[col_idx] if col_idx < len(month_row) else None
        if hasattr(val, "strftime"):
            months.append(val.strftime("%Y-%m"))
        elif isinstance(val, str) and len(val) >= 7:
            months.append(val[:7])
        else:
            months.append(None)

    imported, skipped, flagged = 0, 0, []
    errors = []

    for row in rows[2:]:
        if not row or row[0] is None:
            break

        link_name = str(row[0]).strip()
        bw_gbps   = int(row[3]) if row[3] else 100

        # Parsear los 6 meses (cada mes ocupa 4 columnas: Max, %Max, AVG, %AVG)
        for i, month in enumerate(months):
            if not month:
                continue
            base = 4 + i * 4
            if base + 3 >= len(row):
                continue
            max_gbps = row[base]
            pct_max  = row[base + 1]
            avg_gbps = row[base + 2]
            pct_avg  = row[base + 3]

            # Detectar flags
            row_flags = []
            if bw_gbps >= 200:
                row_flags.append("DOUBLE_LAMBDA")

            # Resolver sitios del link_name
            parts = [p.strip() for p in link_name.replace(" - ", "-").split("-", 1)]
            site_a = _resolve_site(parts[0]) if len(parts) > 0 else None
            site_b = _resolve_site(parts[1]) if len(parts) > 1 else None

            if site_a is None or site_b is None:
                row_flags.append("UNKNOWN_SITE")

            # Buscar lambda correspondiente
            lm = _find_lambda_for_link(site_a, site_b) if site_a and site_b else None

            flags_str = ",".join(row_flags) if row_flags else None

            # Upsert: si ya existe este (month, link_name), actualizar
            existing = LambdaUtilization.query.filter_by(month=month, link_name=link_name).first()
            if existing:
                existing.bw_gbps  = bw_gbps
                existing.max_gbps = max_gbps
                existing.pct_max  = pct_max
                existing.avg_gbps = avg_gbps
                existing.pct_avg  = pct_avg
                existing.flags    = flags_str
                existing.lambda_id = lm.id if lm else None
            else:
                rec = LambdaUtilization(
                    month=month, link_name=link_name, bw_gbps=bw_gbps,
                    max_gbps=max_gbps, pct_max=pct_max,
                    avg_gbps=avg_gbps, pct_avg=pct_avg,
                    flags=flags_str, lambda_id=lm.id if lm else None,
                )
                db.session.add(rec)
                imported += 1

            if row_flags:
                flagged.append({
                    "month": month, "link_name": link_name,
                    "bw_gbps": bw_gbps, "flags": row_flags,
                    "site_a": site_a, "site_b": site_b,
                })

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Error al guardar: {e}"}), 500

    months_loaded = [m for m in months if m]
    return jsonify({
        "status": "ok",
        "months_loaded": months_loaded,
        "records_imported": imported,
        "flagged_rows": flagged,
    })


@bp.get("/lambda-utilization")
def get_lambda_utilization():
    """Retorna todos los registros de utilización agrupados por mes."""
    month_filter = request.args.get("month")
    q = LambdaUtilization.query
    if month_filter:
        q = q.filter_by(month=month_filter)
    records = q.order_by(LambdaUtilization.month, LambdaUtilization.link_name).all()

    # Agrupar por mes
    by_month = {}
    for r in records:
        by_month.setdefault(r.month, []).append(r.to_dict())

    # Lista de meses disponibles
    months = sorted(by_month.keys())

    return jsonify({
        "months": months,
        "data": by_month,
    })
