"""API de reportes — exportación PDF, Excel y CSV."""
import csv
import io
from flask import Blueprint, Response, request
from ..models import Site, Lambda, Segment
from ..extensions import db
from sqlalchemy import func

bp = Blueprint("reports", __name__)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _sites_data():
    return Site.query.order_by(Site.type, Site.name).all()

def _lambdas_data():
    return Lambda.query.order_by(Lambda.name).all()

def _segments_data():
    segs = Segment.query.all()
    segs.sort(key=lambda s: s.usage_count, reverse=True)
    return segs


# ── CSV ───────────────────────────────────────────────────────────────────────

@bp.get("/reports/sites.csv")
def sites_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Nombre", "Tipo", "Región", "Ciudad", "Lat", "Lon"])
    for s in _sites_data():
        writer.writerow([s.id, s.name, s.type, s.region, s.city, s.lat, s.lon])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=sitios.csv"},
    )


@bp.get("/reports/lambdas.csv")
def lambdas_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Nombre", "Color", "Num Lambdas", "Capacidad/Lambda (Gbps)",
                     "Capacidad Total (Gbps)", "Ruta Protección"])
    for l in _lambdas_data():
        writer.writerow([
            l.id, l.name, l.color, l.num_lambdas,
            l.capacity_per_lambda, l.total_capacity_gbps, l.protection_route_name or "",
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=lambdas.csv"},
    )


@bp.get("/reports/segments.csv")
def segments_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Sitio A", "Sitio B", "Fibra", "Proveedor",
                     "Lambdas en uso", "% de capacidad", "Alerta"])
    for s in _segments_data():
        writer.writerow([
            s.id, s.site_a_id, s.site_b_id, s.fiber, s.fiber_provider,
            s.usage_count, f"{s.usage_percent}%", "SÍ" if s.usage_count >= 77 else "No",
        ])
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=segmentos.csv"},
    )


# ── Excel ─────────────────────────────────────────────────────────────────────

@bp.get("/reports/general.xlsx")
def general_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "openpyxl no instalado."}, 500

    wb = Workbook()

    # Hoja 1 — Sitios
    ws_sites = wb.active
    ws_sites.title = "Sitios"
    headers = ["ID", "Nombre", "Tipo", "Región", "Ciudad", "Lat", "Lon"]
    ws_sites.append(headers)
    for cell in ws_sites[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    for s in _sites_data():
        ws_sites.append([s.id, s.name, s.type, s.region, s.city, s.lat, s.lon])

    # Hoja 2 — Lambdas
    ws_lam = wb.create_sheet("Lambdas")
    ws_lam.append(["ID", "Nombre", "Color", "Num Lambdas", "Cap/Lambda (Gbps)",
                   "Cap Total (Gbps)", "Ruta Protección"])
    for cell in ws_lam[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for l in _lambdas_data():
        ws_lam.append([
            l.id, l.name, l.color, l.num_lambdas,
            l.capacity_per_lambda, l.total_capacity_gbps, l.protection_route_name or "",
        ])

    # Hoja 3 — Segmentos
    ws_seg = wb.create_sheet("Segmentos")
    ws_seg.append(["ID", "Sitio A", "Sitio B", "Fibra", "Proveedor",
                   "Lambdas uso", "% capacidad", "Alerta"])
    for cell in ws_seg[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    from openpyxl.styles import PatternFill as PF
    red_fill = PF("solid", fgColor="FF6B6B")
    for seg in _segments_data():
        row = [
            seg.id, seg.site_a_id, seg.site_b_id, seg.fiber, seg.fiber_provider,
            seg.usage_count, seg.usage_percent, "⚠️ ALERTA" if seg.usage_count >= 77 else "",
        ]
        ws_seg.append(row)
        if seg.usage_count >= 77:
            for cell in ws_seg[ws_seg.max_row]:
                cell.fill = red_fill

    # Ajustar anchos
    for ws in [ws_sites, ws_lam, ws_seg]:
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 20

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_dwdm.xlsx"},
    )


# ── PDF ───────────────────────────────────────────────────────────────────────

@bp.get("/reports/general.pdf")
def general_pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
    except ImportError:
        return {"error": "reportlab no instalado."}, 500

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm)
    styles = getSampleStyleSheet()
    story = []

    # Título
    story.append(Paragraph("Reporte DWDM — Red ISP Tx", styles["Title"]))
    story.append(Spacer(1, 0.5*cm))

    # KPIs
    from ..models import Site as SiteM, Lambda as LambdaM, Segment as SegmentM
    kpis = [
        ["Sitios propios", SiteM.query.filter_by(type="own").count()],
        ["Sitios terceros", SiteM.query.filter_by(type="third_party").count()],
        ["Lambdas activas", LambdaM.query.count()],
        ["Segmentos únicos", SegmentM.query.count()],
    ]
    kpi_table = Table([["Métrica", "Valor"]] + kpis, colWidths=[10*cm, 5*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF5FB")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 1*cm))

    # Segmentos
    story.append(Paragraph("Segmentos por uso", styles["Heading2"]))
    seg_data = [["ID", "Sitio A", "Sitio B", "Fibra", "Proveedor", "Lambdas", "% Cap"]]
    for seg in _segments_data()[:30]:  # Top 30
        seg_data.append([
            str(seg.id), seg.site_a_id, seg.site_b_id, seg.fiber,
            seg.fiber_provider or "-", str(seg.usage_count), f"{seg.usage_percent}%",
        ])
    seg_table = Table(seg_data, colWidths=[1.5*cm, 5*cm, 5*cm, 2.5*cm, 4*cm, 2.5*cm, 2.5*cm])
    seg_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF5FB")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(seg_table)

    doc.build(story)
    out.seek(0)
    return Response(
        out.read(),
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_dwdm.pdf"},
    )
